#%% *args形参的使用: 当传入多个位置参数，这多个位置参数会自动组成一个元组，然后就可以遍历这个元组中的参数
def fun(*args):
    print(type(args))
    print(args)
    for item in args:
        print(item)
fun(2, 'alex', [3])


#%% 在列表、元素、字典、集合和numpy的数组前加上星号 * 的含义: 将数据结构中元素都被分成一个一个的独立元素
List = ['a', 2, 3]
Tuple = ('b', 'c', 5)
Dict = {'name': 'Ada', 'age': 23}

print(List)
print(Tuple)
print(Dict)

print(*List)
print(*Tuple)
print(*Dict)

def print_item(*args):
    print(type(args))
    print(args)
    for item in args:
        print(item)

List = ['apple', 23, 'orange', 9]
print_item(List) # list这个元组只有一个元素
print_item(*List) # 加上 * 之后， 直接把列表里面的元素取出来放到一个元组中， 然后直接遍历



#%% **kwargs形参的使用: 当传入多个关键字参数，这多个位置参数会自动组成一个字典，然后就可以遍历这个字典中的参数
def fun(**kwargs):
    print(type(kwargs))
    print(kwargs)
    for key in kwargs.keys():
        print(key, kwargs[key])

fun(name='Alvin', age=23, things=[1, 2, 'a'])






#%% 创建文件夹
import os
def make_directory(makepath):
    makepath = makepath.strip() # 去除首位空格
    makepath = makepath.rstrip("\\") # 去除尾部 \ 符号
    isExists = os.path.exists(makepath) # 判断路径是否存在：存在为True，不存在为False
    # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        os.makedirs(makepath) # 创建目录操作函数
        print(makepath + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(makepath + ' 目录已存在')
        return False

makepath = "C:\\Users\\lenovo\\Desktop\\mathematicsModelingPython.git\\mathematicsModelingPython\\data of data_preprocessing\\"
make_directory(makepath)



#%% 写入txt文本
def txt_create(make_path, file_name, information):
    # make_path为新创建的txt文件的存放路径
    full_path = make_path + file_name + '.txt'  # 也可以创建一个.doc的word文档
    with open(full_path, "w") as file:
        file.write("这是个\n测试！")  # 自带文件关闭功能，不需要再写f.close()
    #上面with open() as f的两行语句也可以用下面这三行替换
    # file = open(full_path, 'w')
    # file.write(information)
    # file.close()

make_path = "C:\\Users\\lenovo\\Desktop\\mathematicsModelingPython.git\\mathematicsModelingPython\\data of data_preprocessing\\"
file_name = 'example_of_txt_create'
information = 'My Hello world!! !'
txt_create(make_path, file_name, information)



#%% 读取txt文本
filepath = "data of data_preprocessing\\example_of_txt_data.txt"

with open(filepath, "r") as file1:  # 打开文件
    data1 = file1.read()  # read() 一次性读全部内容 ： 以字符串的形式返回结果
    data2 = file1.readline() # readline() 读取第一行内容 ： 以字符串的形式返回结果
    data3 = file1.readlines() # readlines() 列表 ： 读取文本所有内容，并且以数列的格式返回结果，一般配合for in使用
    print(data1)
    print(data2)
    print(data3)







#%% 创建Excel文件
import openpyxl # 用openpyxl模块中的Workbook来创建Excel文件
import datetime # 用datetime模块来读取时间

workbook = openpyxl.Workbook()
worksheet = workbook.active # 激活 workbook
workbook.active.title = 'New Title'
worksheet['A1'] = 42 # 数据可以直接分配到单元格中
worksheet.append([1,2,3]) # 可以附加行，从第一列开始附加
worksheet['A3'] = datetime.datetime.now().strftime("%Y-%m-%d") # 用strftime格式化时间, 默认为计算机当前的时间，
# %Y对应年，%m对应月，%d对应日，还有时、分、秒，格式为strftime("%Y-%m-%d %H:%M:%S")，其中%H为24小时制，可改为$l为12小时制
for i in range(10):
    worksheet['D%d' % (i+1)] = i + 1 # 向D1到D10添加数据(range(10)=[0,1,2,3,4,5,6,7,8,9])
worksheet['E2'] = '=SUM(A:A)' # 向excel表中输入表达式
workbook.save("data of data_preprocessing\\example_of_excel_create.xlsx") # 保存文件





#%% 读取Excel文件
import openpyxl #用openpyxl模块中的load_workbook来读取Excel文件

workbook1 = openpyxl.load_workbook('data of data_preprocessing\\example_of_excel_data.xlsx')
# print(workbook1.sheetnames) # 用sheetnames输出所有sheet的名称
worksheet1 = workbook1['profits'] # 选取workbook1中名称为profits的sheet
# for row in worksheet.values:
#     print(row)
dictionary = {} # 字典可以实现一个行业对应一个净利润
# iter_rows函数用于行的循环, min_row=2表示从第2行开始，max_col=5表示到第5列结束
for row in worksheet1.iter_rows(min_row=2, max_col=5, max_row=len(worksheet1['D'])):
    # print(row[3].value)
    dictionary[row[3].value] = dictionary.get(row[3].value, 0) + row[4].value # get函数表示如果找到了row[3.value则返回row[3].value, 找不到则返回0
# print(dictionary)
use_range = worksheet1['H1':'I3602']
for i in use_range:
    if dictionary.get(i[0].value, 0) != 0:
        i[1].value = dictionary[i[0].value]
workbook1.save("data of data_preprocessing\\example_of_excel_data_finished.xlsx") # 将上述dictionary写入example_of_excel_data的I列

workbook2 = openpyxl.load_workbook('data of data_preprocessing\\PE_data_original.xlsx')
# print(workbook2.sheetnames) # 用sheetnames输出所有sheet的名称
for i in range(len(workbook2.sheetnames)):
    worksheet2 = workbook2.worksheets[i]
    worksheet2.title = 'a' + str(i)
workbook2.save('data of data_preprocessing\\PE_data_revised.xlsx')



#%% 读取写入文件Example
filepath = 'data of data_preprocessing\\szzs.txt'
date = []
percent = []
with open(filepath, "r") as file1:  # 读取txt文件
    for line in file1:
        a,b = line.split() # line.split会自动按空格劈开，szzs.txt被劈成两部分
        date.append(a)
        percent.append(b)
new_percent = []
data = {}
for i in percent:
    new_percent.append(float(i)) # 将percent内的文本数据变成浮点型数据(即将字符串格式string变成数字格式)
for i in range(len(date)):
    data[date[i]] = new_percent[i]
count = 0
results = {}
for k,v in data.items():# 将连续上涨的总周数和对应的被终结日期存入results字典
    if v > 0:
        count += 1
    else:
        date_list = []
        date_list.append(k)
        results[count] = results.get(count, []) + date_list
        count = 0
with open('data of data_preprocessing\\上证指数连涨统计数据.txt', 'w') as file2:
    for k, v in results.items():
        if k >= 6:
            file2.write('连续{}周的情况对应的被终结日期是: {}\n'.format(k, v))




#%% 读取写入文件中用到的缩写字母含义汇总
# r : 读取文件，若文件不存在则会报错
# w: 写入文件，若文件不存在则会先创建再写入，会覆盖原文件
# a : 写入文件，若文件不存在则会先创建再写入，但不会覆盖原文件，而是追加在文件末尾
# rb,wb： 分别于r,w类似，但是用于读写二进制文件
# r+ : 可读、可写，文件不存在也会报错，写操作时会覆盖
# w+ : 可读，可写，文件不存在先创建，会覆盖
# a+ : 可读、可写，文件不存在先创建，不会覆盖，追加在末尾



